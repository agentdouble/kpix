import csv
import io
import logging
from datetime import date, datetime
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from kpix_backend.api.deps_resources import get_kpi_or_404
from kpix_backend.core.deps import get_current_user
from kpix_backend.core.db import get_session
from kpix_backend.core.enums import ImportStatus, ImportType
from kpix_backend.core.kpi_logic import compute_status
from kpix_backend.models import DataImportJob, Kpi, KpiValue, User
from kpix_backend.schemas.import_job import ImportJobPublic, ImportResult

router = APIRouter(prefix="/imports", tags=["imports"])
logger = logging.getLogger("kpix")

REQUIRED_COLUMNS = {"kpi_id", "period_start", "value"}


def _parse_date(value: str | datetime | date) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError as exc:
            raise ValueError(f"Invalid date format: {value}") from exc
    raise ValueError(f"Unsupported date type: {type(value)}")


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text_stream = io.StringIO(content.decode("utf-8"))
    reader = csv.DictReader(text_stream)
    rows = [row for row in reader]
    if not rows:
        raise ValueError("CSV file is empty")
    missing = REQUIRED_COLUMNS - set(reader.fieldnames or [])
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
    return rows


def _parse_excel(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() for h in next(rows_iter)]
    except StopIteration as exc:
        raise ValueError("Excel file is empty") from exc
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
    rows: list[dict[str, str]] = []
    for row in rows_iter:
        row_dict = {header: row[idx] for idx, header in enumerate(headers)}
        rows.append(row_dict)
    if not rows:
        raise ValueError("Excel file contains only headers")
    return rows


async def _ingest_rows(
    rows: list[dict[str, str]],
    session: AsyncSession,
    current_user: User,
) -> int:
    ingested = 0
    for row in rows:
        try:
            kpi_id = uuid.UUID(str(row["kpi_id"]))
            period_start = _parse_date(row["period_start"])
            period_end_raw = row.get("period_end") or row.get("period_end".capitalize())
            period_end = _parse_date(period_end_raw) if period_end_raw else period_start
            value = float(row["value"])
            comment = row.get("comment")
        except Exception as exc:
            raise ValueError(f"Invalid row data: {row}") from exc

        kpi: Kpi = await get_kpi_or_404(session, kpi_id, current_user.organization_id)
        status_value = compute_status(kpi.direction, float(kpi.threshold_green), float(kpi.threshold_orange), value)
        kpi_value = KpiValue(
            kpi_id=kpi.id,
            organization_id=current_user.organization_id,
            period_start=period_start,
            period_end=period_end,
            value=value,
            status=status_value,
            comment=comment,
        )
        session.add(kpi_value)
        ingested += 1
    return ingested


@router.post("/kpi-values", response_model=ImportResult, status_code=status.HTTP_201_CREATED)
async def import_kpi_values(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ImportResult:
    job = DataImportJob(
        organization_id=current_user.organization_id,
        type=ImportType.EXCEL,
        status=ImportStatus.PENDING,
        created_by=current_user.id,
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)

    try:
        job.status = ImportStatus.RUNNING
        await session.commit()

        content = await file.read()
        filename = (file.filename or "").lower()
        if filename.endswith(".csv"):
            rows = _parse_csv(content)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_excel(content)
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file type")

        ingested_count = await _ingest_rows(rows, session, current_user)
        job.status = ImportStatus.SUCCESS
        job.error_message = None
        await session.commit()
        logger.info(
            "import_completed",
            extra={
                "job_id": str(job.id),
                "organization_id": str(current_user.organization_id),
                "ingested": ingested_count,
            },
        )
        return ImportResult(job_id=job.id, ingested=ingested_count, failed=0, errors=[])
    except HTTPException as exc:
        await session.rollback()
        job.status = ImportStatus.FAILED
        job.error_message = str(exc.detail)
        await session.commit()
        logger.exception(
            "import_failed",
            extra={"job_id": str(job.id), "organization_id": str(current_user.organization_id)},
        )
        raise
    except (ValueError, IntegrityError) as exc:
        await session.rollback()
        job.status = ImportStatus.FAILED
        job.error_message = str(exc)
        await session.commit()
        logger.exception(
            "import_failed",
            extra={"job_id": str(job.id), "organization_id": str(current_user.organization_id)},
        )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    finally:
        await file.close()


@router.get("/jobs", response_model=list[ImportJobPublic])
async def list_import_jobs(
    session: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> list[ImportJobPublic]:
    result = await session.execute(
        select(DataImportJob).where(DataImportJob.organization_id == current_user.organization_id)
    )
    jobs = result.scalars().all()
    return [ImportJobPublic.model_validate(job) for job in jobs]


@router.get("/jobs/{job_id}", response_model=ImportJobPublic)
async def get_import_job(
    job_id: uuid.UUID,
    session: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ImportJobPublic:
    job = await session.get(DataImportJob, job_id)
    if not job or job.organization_id != current_user.organization_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import job not found")
    return ImportJobPublic.model_validate(job)
